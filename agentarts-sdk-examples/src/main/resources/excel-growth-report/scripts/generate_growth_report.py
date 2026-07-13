#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Business Growth Report Generator - Calculate YoY growth and output Excel
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.utils.dataframe import dataframe_to_rows


def inspect_data(df, n_rows=5):
    """
    Preview first n rows of data to understand structure.
    Returns as Markdown table for LLM readability.

    Args:
        df: pandas DataFrame
        n_rows: number of rows to preview

    Returns:
        str: Markdown formatted table
    """
    return df.head(n_rows).to_markdown(index=False)


def validate_columns(df):
    """
    Validate that required columns exist.

    Args:
        df: pandas DataFrame

    Returns:
        tuple: (is_valid, error_message)
    """
    required = ['日期', 'metric', 'value']
    # Also check English column names
    required_en = ['date', 'metric', 'value']

    cols = df.columns.tolist()

    # Check Chinese columns
    if all(col in cols for col in required):
        return True, None

    # Check English columns
    if all(col in cols for col in required_en):
        # Rename to Chinese for consistency
        df.rename(columns={
            'date': '日期',
            'metric': '指标',
            'value': '数值'
        }, inplace=True)
        return True, None

    return False, f"Missing required columns. Expected: {required} or {required_en}"


def calculate_growth(df):
    """
    Calculate year-over-year growth rates.

    Args:
        df: pandas DataFrame with columns: 日期, 指标, 数值

    Returns:
        pandas DataFrame: pivoted data with growth rates
    """
    # Pivot data: rows = dates, columns = metrics
    pivot = df.pivot_table(index='日期', columns='指标', values='数值')
    pivot = pivot.sort_index()

    # Calculate YoY growth (12-month period)
    pivot['dau_yoy'] = pivot['dau'].pct_change(periods=12) * 100
    pivot['revenue_yoy'] = pivot['revenue'].pct_change(periods=12) * 100

    return pivot


def format_excel_sheet(worksheet, pivot_df):
    """
    Apply formatting to Excel worksheet.

    Args:
        worksheet: openpyxl worksheet object
        pivot_df: pandas DataFrame with growth data
    """
    # Find column indices for growth columns
    headers = [cell.value for cell in worksheet[1]]

    dau_col_idx = None
    revenue_col_idx = None

    for idx, header in enumerate(headers, start=1):
        if header == 'dau_yoy':
            dau_col_idx = idx
        elif header == 'revenue_yoy':
            revenue_col_idx = idx

    # Apply conditional formatting to growth columns
    for row in range(2, worksheet.max_row + 1):
        if dau_col_idx:
            cell = worksheet.cell(row=row, column=dau_col_idx)
            if cell.value is not None:
                cell.number_format = '0.00%'
                if cell.value < 0:
                    cell.font = Font(color='FF0000')  # Red for negative
                elif cell.value > 0:
                    cell.font = Font(color='008000')  # Green for positive

        if revenue_col_idx:
            cell = worksheet.cell(row=row, column=revenue_col_idx)
            if cell.value is not None:
                cell.number_format = '0.00%'
                if cell.value < 0:
                    cell.font = Font(color='FF0000')
                elif cell.value > 0:
                    cell.font = Font(color='008000')

    # Auto-adjust column widths
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    worksheet.freeze_panes = 'A2'


def add_charts_to_dashboard(workbook, dashboard_sheet, pivot_df):
    """
    Add charts to the summary dashboard sheet.

    Args:
        workbook: openpyxl workbook object
        dashboard_sheet: openpyxl worksheet object
        pivot_df: pandas DataFrame with growth data
    """
    # Create data reference from the '数据与增长' sheet
    data_sheet = workbook['数据与增长']

    # Bar chart: DAU YoY Growth
    chart = BarChart()
    chart.title = "DAU Year-over-Year Growth"
    chart.x_axis.title = "Period"
    chart.y_axis.title = "Growth Rate (%)"

    # Reference data
    data = Reference(data_sheet, min_col=3, min_row=1, max_row=len(pivot_df)+1)
    categories = Reference(data_sheet, min_col=1, min_row=2, max_row=len(pivot_df)+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.width = 15
    chart.height = 10

    dashboard_sheet.add_chart(chart, "F2")


def generate_growth_report(df, output_dir="./"):
    """
    Main function: process data and generate Excel report.

    Args:
        df: pandas DataFrame with columns: 日期, 指标, 数值
        output_dir: directory to save the report

    Returns:
        dict: result with status, file path, and error count
    """
    try:
        # Validate columns
        is_valid, error_msg = validate_columns(df)
        if not is_valid:
            return {
                "status": "error",
                "message": error_msg
            }

        # Convert date column
        df['日期'] = pd.to_datetime(df['日期'])

        # Calculate growth
        pivot_df = calculate_growth(df)

        # Generate summary statistics
        summary = {
            'Period': f"{pivot_df.index.min().strftime('%Y-%m')} to {pivot_df.index.max().strftime('%Y-%m')}",
            'Total DAU': int(pivot_df['dau'].sum()),
            'Total Revenue': int(pivot_df['revenue'].sum()),
            'Avg DAU Growth': f"{pivot_df['dau_yoy'].mean():.2f}%",
            'Avg Revenue Growth': f"{pivot_df['revenue_yoy'].mean():.2f}%"
        }

        # Create Excel file
        filename = f"growth_report_{datetime.now().strftime('%Y%m%d')}.xlsx"
        output_path = Path(output_dir) / filename
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Sheet 1: Data & Growth
            pivot_df.to_excel(writer, sheet_name='数据与增长')

            # Sheet 2: Summary Dashboard
            summary_df = pd.DataFrame([summary])
            summary_df.to_excel(writer, sheet_name='汇总看板', index=False)

            # Get workbook and apply formatting
            workbook = writer.book

            # Format data sheet
            data_sheet = writer.sheets['数据与增长']
            format_excel_sheet(data_sheet, pivot_df)

            # Add charts to dashboard
            dashboard_sheet = writer.sheets['汇总看板']
            add_charts_to_dashboard(workbook, dashboard_sheet, pivot_df)

        return {
            "status": "success",
            "file": str(output_path),
            "errors": 0,
            "summary": summary
        }

    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }


def validate_report(file_path):
    """
    Validate Excel file for formula errors.

    Args:
        file_path: path to Excel file

    Returns:
        dict: validation result
    """
    try:
        wb = load_workbook(file_path, data_only=False)
        error_cells = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f':  # Formula cell
                        # Check if formula has error
                        if cell.value and isinstance(cell.value, str):
                            error_keywords = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?']
                            if any(keyword in cell.value for keyword in error_keywords):
                                error_cells.append({
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'formula': cell.value
                                })

        return {
            "valid": len(error_cells) == 0,
            "error_cells": error_cells
        }

    except Exception as e:
        return {
            "valid": False,
            "error": str(e)
        }


# Command line entry point
if __name__ == "__main__":
    import sys
    import argparse

    parser = argparse.ArgumentParser(description='Generate growth report from business data')
    parser.add_argument('file', help='Input data file (CSV or Excel)')
    parser.add_argument('--output', '-o', default='./', help='Output directory')
    parser.add_argument('--inspect', '-i', action='store_true', help='Inspect data only')
    parser.add_argument('--validate', '-v', help='Validate existing report file')

    args = parser.parse_args()

    if args.validate:
        result = validate_report(args.validate)
        print(json.dumps(result, indent=2, ensure_ascii=False))
        sys.exit(0)

    # Read input file
    if args.file.endswith('.csv'):
        df = pd.read_csv(args.file)
    else:
        df = pd.read_excel(args.file)

    if args.inspect:
        print(inspect_data(df))
        sys.exit(0)

    result = generate_growth_report(df, args.output)
    print(json.dumps(result, indent=2, ensure_ascii=False))